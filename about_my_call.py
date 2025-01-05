import openpyxl


def check_my_calls(call_roaster,house_officer_name):
    # Specify the Excel file path to be edited
    excel_file_path = call_roaster

    # Load the existing Excel workbook
    workbook = openpyxl.load_workbook(excel_file_path)

    # Select the sheet you want to edit (assuming the first sheet for simplicity)
    sheet = workbook.active

    row_with_call = []
    col_with_call = []

    # Iterate through all rows and columns
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):

        for cell in row:
            # Example: Change values based on some condition
            
            cell_value = cell.value
                       
            if type(cell_value) is str and cell_value.count(house_officer_name) > 0:
                row_with_call.append(cell.row)
                col_with_call.append(cell.column)
                pass
        for cell in row:
            if cell.row == 1 and cell.column == 1:
                cell.value = house_officer_name

            elif cell.row in row_with_call or cell.row == 1:
                pass
                
            else:
                cell.value = ''

       
        
    print(row_with_call)       
    # # Save the changes to the Excel file
    filename = input('new_filename.xlsx:   ')
    full_filename = f'{filename}.xlsx'

    workbook.save(full_filename)



call_roaster= input('Enter the filename(including the file extension):  ')  # input('file name (eg. callroaster.xlsx):')

ho_name = ''
while ho_name != 'QUIT':
    ho_name = input('HO name: ').upper()
   
    check_my_calls(call_roaster,ho_name)
    